from fastapi import APIRouter, Depends, Query, HTTPException, Response
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import func, and_, extract, text
from typing import List, Dict, Any, Optional
from datetime import datetime, timedelta, date
from decimal import Decimal
import csv
import io
import os
import zipfile
import json
from ..database import get_db
from .. import crud
from ..models.transaction import Transaction
from ..models.category import Category
from ..models.budget import Budget
from ..models.account import Account
from ..models.user import User
from ..models.tag import Tag
from sqlalchemy import select

router = APIRouter()


@router.get("/summary")
async def get_summary(
    db: AsyncSession = Depends(get_db),
    start_date: str = Query(..., description="Start date (YYYY-MM-DD)"),
    end_date: str = Query(..., description="End date (YYYY-MM-DD)"),
    user_id: int = Query(..., description="User ID")
):
    """收支概览统计（总收入、总支出、净收入）"""
    start_dt = datetime.strptime(start_date, "%Y-%m-%d")
    end_dt = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)

    # 获取收入总额
    income_stmt = select(func.sum(Transaction.amount)).where(
        and_(
            Transaction.user_id == user_id,
            Transaction.transaction_type == 'income',
            Transaction.date >= start_dt,
            Transaction.date <= end_dt,
            Transaction.is_active == True
        )
    )
    income_result = await db.execute(income_stmt)
    total_income = income_result.scalar() or 0

    # 获取支出总额
    expense_stmt = select(func.sum(Transaction.amount)).where(
        and_(
            Transaction.user_id == user_id,
            Transaction.transaction_type == 'expense',
            Transaction.date >= start_dt,
            Transaction.date <= end_dt,
            Transaction.is_active == True
        )
    )
    expense_result = await db.execute(expense_stmt)
    total_expense = expense_result.scalar() or 0

    # 计算净收入
    net_income = total_income - total_expense

    return {
        "total_income": float(total_income),
        "total_expense": float(total_expense),
        "net_income": float(net_income),
        "date_range": {
            "start": start_date,
            "end": end_date
        }
    }


@router.get("/category-analysis")
async def get_category_analysis(
    db: AsyncSession = Depends(get_db),
    start_date: str = Query(..., description="Start date (YYYY-MM-DD)"),
    end_date: str = Query(..., description="End date (YYYY-MM-DD)"),
    user_id: int = Query(..., description="User ID"),
    transaction_type: str = Query("expense", description="Type of transaction (income/expense)")
):
    """分类支出占比（饼图数据）"""
    start_dt = datetime.strptime(start_date, "%Y-%m-%d")
    end_dt = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)

    stmt = select(
        func.sum(Transaction.amount).label('total'),
        Transaction.category_id,
        Category.name.label('category_name')
    ).select_from(
        Transaction.__table__.join(Category.__table__)
    ).where(
        and_(
            Transaction.user_id == user_id,
            Transaction.transaction_type == transaction_type,
            Transaction.date >= start_dt,
            Transaction.date <= end_dt,
            Transaction.is_active == True,
            Transaction.category_id == Category.id
        )
    ).group_by(Transaction.category_id, Category.name)

    result = await db.execute(stmt)
    rows = result.fetchall()

    total_amount = sum(float(row[0] or 0) for row in rows)

    category_data = []
    for row in rows:
        total = float(row[0] or 0)
        category_id = row[1]
        category_name = row[2]

        percentage = round((total / total_amount * 100) if total_amount > 0 else 0, 2)
        category_data.append({
            "category_id": category_id,
            "category_name": category_name,
            "amount": total,
            "percentage": percentage
        })

    return {
        "total_amount": total_amount,
        "categories": category_data,
        "date_range": {
            "start": start_date,
            "end": end_date
        },
        "transaction_type": transaction_type
    }


@router.get("/trend-analysis")
async def get_trend_analysis(
    db: AsyncSession = Depends(get_db),
    start_date: str = Query(..., description="Start date (YYYY-MM-DD)"),
    end_date: str = Query(..., description="End date (YYYY-MM-DD)"),
    user_id: int = Query(..., description="User ID"),
    group_by: str = Query("day", description="Group by (day, week, month)")
):
    """收支趋势（折线图数据）- MySQL compatible"""
    start_dt = datetime.strptime(start_date, "%Y-%m-%d")
    end_dt = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)

    # MySQL DATE_FORMAT for grouping
    if group_by == 'day':
        fmt = '%Y-%m-%d'
    elif group_by == 'week':
        fmt = '%Y-%u'  # ISO week number
    else:
        fmt = '%Y-%m'

    # 收入趋势查询
    income_stmt = select(
        func.date_format(Transaction.date, fmt).label('date_group'),
        func.sum(Transaction.amount).label('total')
    ).where(
        and_(
            Transaction.user_id == user_id,
            Transaction.transaction_type == 'income',
            Transaction.date >= start_dt,
            Transaction.date <= end_dt,
            Transaction.is_active == True
        )
    ).group_by(func.date_format(Transaction.date, fmt)).order_by(
        func.date_format(Transaction.date, fmt)
    )

    income_result = await db.execute(income_stmt)
    income_rows = income_result.fetchall()

    # 支出趋势查询
    expense_stmt = select(
        func.date_format(Transaction.date, fmt).label('date_group'),
        func.sum(Transaction.amount).label('total')
    ).where(
        and_(
            Transaction.user_id == user_id,
            Transaction.transaction_type == 'expense',
            Transaction.date >= start_dt,
            Transaction.date <= end_dt,
            Transaction.is_active == True
        )
    ).group_by(func.date_format(Transaction.date, fmt)).order_by(
        func.date_format(Transaction.date, fmt)
    )

    expense_result = await db.execute(expense_stmt)
    expense_rows = expense_result.fetchall()

    # 转换为字典
    income_dict = {str(row.date_group): float(row.total or 0) for row in income_rows}
    expense_dict = {str(row.date_group): float(row.total or 0) for row in expense_rows}

    # 生成时间序列
    trend_data = []
    current = start_dt.replace(hour=0, minute=0, second=0, microsecond=0)

    if group_by == 'day':
        while current <= end_dt:
            date_key = current.strftime('%Y-%m-%d')
            income_val = income_dict.get(date_key, 0)
            expense_val = expense_dict.get(date_key, 0)
            trend_data.append({
                "date": date_key,
                "income": income_val,
                "expense": expense_val,
                "net": income_val - expense_val
            })
            current += timedelta(days=1)
    elif group_by == 'week':
        # 从周一开始
        current = current - timedelta(days=current.weekday())
        while current <= end_dt:
            date_key = current.strftime('%Y-%u')
            income_val = income_dict.get(date_key, 0)
            expense_val = expense_dict.get(date_key, 0)
            trend_data.append({
                "date": date_key,
                "income": income_val,
                "expense": expense_val,
                "net": income_val - expense_val
            })
            current += timedelta(weeks=1)
    else:  # month
        while (current.year < end_dt.year) or (current.year == end_dt.year and current.month <= end_dt.month):
            date_key = current.strftime('%Y-%m')
            income_val = income_dict.get(date_key, 0)
            expense_val = expense_dict.get(date_key, 0)
            trend_data.append({
                "date": date_key,
                "income": income_val,
                "expense": expense_val,
                "net": income_val - expense_val
            })
            if current.month == 12:
                current = current.replace(year=current.year + 1, month=1)
            else:
                current = current.replace(month=current.month + 1)

    return {
        "trend_data": trend_data,
        "group_by": group_by,
        "date_range": {
            "start": start_date,
            "end": end_date
        }
    }


@router.get("/monthly-report")
async def get_monthly_report(
    db: AsyncSession = Depends(get_db),
    year: int = Query(..., description="Year for the report"),
    month: int = Query(..., description="Month for the report"),
    user_id: int = Query(..., description="User ID")
):
    """月度报表汇总"""
    import calendar
    start_date = datetime(year, month, 1)
    _, last_day = calendar.monthrange(year, month)
    end_date = datetime(year, month, last_day, 23, 59, 59)

    start_str = start_date.strftime("%Y-%m-%d")
    end_str = end_date.strftime("%Y-%m-%d")

    # 获取月度摘要
    summary = await get_summary(db, start_str, end_str, user_id)

    # 获取分类分析
    category_analysis = await get_category_analysis(db, start_str, end_str, user_id, "expense")

    # 获取预算执行情况
    budgets = await crud.budget.get_monthly_budgets(db, user_id, year, month)

    budget_performance = []
    for budget in budgets:
        updated_budget = await crud.budget.update_budget_spent_amount(db, budget.id)
        budget_performance.append({
            "id": updated_budget.id,
            "name": updated_budget.name,
            "category_name": updated_budget.category.name if updated_budget.category else "Unknown",
            "budget_amount": float(updated_budget.amount),
            "spent_amount": float(updated_budget.spent_amount),
            "percentage_used": updated_budget.spent_percentage,
            "is_over_spent": updated_budget.is_over_spent
        })

    return {
        "summary": summary,
        "category_analysis": category_analysis,
        "budget_performance": budget_performance,
        "report_period": {
            "year": year,
            "month": month
        }
    }


@router.get("/budget-performance")
async def get_detailed_budget_performance(
    db: AsyncSession = Depends(get_db),
    start_date: str = Query(..., description="Start date (YYYY-MM-DD)"),
    end_date: str = Query(..., description="End date (YYYY-MM-DD)"),
    user_id: int = Query(..., description="User ID")
):
    """预算执行情况详细分析"""
    start_dt = datetime.strptime(start_date, "%Y-%m-%d")
    end_dt = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)

    stmt = select(Budget).where(
        and_(
            Budget.user_id == user_id,
            Budget.period_start <= end_dt,
            Budget.period_end >= start_dt,
            Budget.is_active == True
        )
    )

    budgets_result = await db.execute(stmt)
    budgets = budgets_result.scalars().all()

    budget_details = []
    for budget in budgets:
        updated_budget = await crud.budget.update_budget_spent_amount(db, budget.id)
        budget_details.append({
            "id": updated_budget.id,
            "name": updated_budget.name,
            "category_name": updated_budget.category.name if updated_budget.category else "Unknown",
            "budget_amount": float(updated_budget.amount),
            "spent_amount": float(updated_budget.spent_amount),
            "percentage_used": updated_budget.spent_percentage,
            "is_over_spent": updated_budget.is_over_spent,
            "period_start": updated_budget.period_start.isoformat() if updated_budget.period_start else None,
            "period_end": updated_budget.period_end.isoformat() if updated_budget.period_end else None
        })

    return {
        "budgets": budget_details,
        "date_range": {
            "start": start_date,
            "end": end_date
        }
    }


# ============================================================
# Phase 5: Data Export (CSV / Excel)
# ============================================================

@router.get("/export/csv")
async def export_transactions_csv(
    db: AsyncSession = Depends(get_db),
    start_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    user_id: int = Query(..., description="User ID"),
    transaction_type: Optional[str] = Query(None, description="income/expense/transfer")
):
    """导出交易记录为 CSV 文件"""
    stmt = (
        select(
            Transaction.id,
            Transaction.date,
            Transaction.amount,
            Transaction.transaction_type,
            Transaction.description,
            Category.name.label('category_name'),
            Account.name.label('account_name'),
            func.group_concat(Tag.name).label('tag_names')
        )
        .outerjoin(Category, Transaction.category_id == Category.id)
        .outerjoin(Account, Transaction.account_id == Account.id)
        .outerjoin(Transaction.tags)
        .where(Transaction.user_id == user_id)
    )

    if start_date:
        stmt = stmt.where(Transaction.date >= datetime.strptime(start_date, "%Y-%m-%d"))
    if end_date:
        stmt = stmt.where(Transaction.date <= datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59))
    if transaction_type:
        stmt = stmt.where(Transaction.transaction_type == transaction_type)

    stmt = stmt.group_by(Transaction.id).order_by(Transaction.date.desc())

    result = await db.execute(stmt)
    rows = result.fetchall()

    # 构建 CSV
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "日期", "金额", "类型", "描述", "分类", "账户", "标签"])

    for row in rows:
        writer.writerow([
            row[0],
            row[1].strftime("%Y-%m-%d %H:%M:%S") if row[1] else "",
            float(row[2]) if row[2] else 0,
            row[3],
            row[4] or "",
            row[5] or "",
            row[6] or "",
            row[7] or ""
        ])

    output.seek(0)
    csv_content = output.getvalue()

    return Response(
        content=csv_content.encode('utf-8-sig'),  # BOM for Excel Chinese support
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="transactions_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'}
    )


@router.get("/export/excel")
async def export_transactions_excel(
    db: AsyncSession = Depends(get_db),
    start_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    user_id: int = Query(..., description="User ID"),
    transaction_type: Optional[str] = Query(None, description="income/expense/transfer")
):
    """导出交易记录为 Excel 文件"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    stmt = (
        select(
            Transaction.id,
            Transaction.date,
            Transaction.amount,
            Transaction.transaction_type,
            Transaction.description,
            Category.name.label('category_name'),
            Account.name.label('account_name'),
            func.group_concat(Tag.name).label('tag_names')
        )
        .outerjoin(Category, Transaction.category_id == Category.id)
        .outerjoin(Account, Transaction.account_id == Account.id)
        .outerjoin(Transaction.tags)
        .where(Transaction.user_id == user_id)
    )

    if start_date:
        stmt = stmt.where(Transaction.date >= datetime.strptime(start_date, "%Y-%m-%d"))
    if end_date:
        stmt = stmt.where(Transaction.date <= datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59))
    if transaction_type:
        stmt = stmt.where(Transaction.transaction_type == transaction_type)

    stmt = stmt.group_by(Transaction.id).order_by(Transaction.date.desc())

    result = await db.execute(stmt)
    rows = result.fetchall()

    # 获取汇总数据
    summary = await get_summary(
        db,
        start_date or "2020-01-01",
        end_date or datetime.now().strftime("%Y-%m-%d"),
        user_id
    )

    # 创建 Excel 工作簿
    wb = openpyxl.Workbook()

    # --- Sheet 1: 交易记录 ---
    ws = wb.active
    ws.title = "交易记录"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = ["ID", "日期", "金额", "类型", "描述", "分类", "账户", "标签"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, row in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=row[0]).border = thin_border
        date_cell = ws.cell(row=row_idx, column=2, value=row[1].strftime("%Y-%m-%d %H:%M:%S") if row[1] else "")
        date_cell.border = thin_border
        amount_cell = ws.cell(row=row_idx, column=3, value=float(row[2]) if row[2] else 0)
        amount_cell.border = thin_border
        amount_cell.number_format = '#,##0.00'
        ws.cell(row=row_idx, column=4, value=row[3]).border = thin_border
        ws.cell(row=row_idx, column=5, value=row[4] or "").border = thin_border
        ws.cell(row=row_idx, column=6, value=row[5] or "").border = thin_border
        ws.cell(row=row_idx, column=7, value=row[6] or "").border = thin_border
        ws.cell(row=row_idx, column=8, value=row[7] or "").border = thin_border

    # 自动调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = min(max_length + 4, 40)

    # --- Sheet 2: 汇总 ---
    ws2 = wb.create_sheet(title="汇总")
    ws2.cell(row=1, column=1, value="总收入").font = Font(bold=True)
    ws2.cell(row=1, column=2, value=summary["total_income"]).number_format = '#,##0.00'
    ws2.cell(row=2, column=1, value="总支出").font = Font(bold=True)
    ws2.cell(row=2, column=2, value=summary["total_expense"]).number_format = '#,##0.00'
    ws2.cell(row=3, column=1, value="净收入").font = Font(bold=True)
    ws2.cell(row=3, column=2, value=summary["net_income"]).number_format = '#,##0.00'

    # 写入内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="transactions_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'}
    )


# ============================================================
# Phase 5: Data Backup & Restore
# ============================================================

BACKUP_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "backups")


def _ensure_backup_dir():
    os.makedirs(BACKUP_DIR, exist_ok=True)


@router.post("/backup")
async def create_backup(
    db: AsyncSession = Depends(get_db),
    user_id: int = Query(..., description="User ID")
):
    """创建用户数据备份（JSON 格式）"""
    _ensure_backup_dir()

    # 导出该用户的所有数据
    # 1. 账户
    accounts_result = await db.execute(
        select(Account).where(Account.user_id == user_id)
    )
    accounts = [
        {
            "id": a.id, "name": a.name, "account_type": a.account_type,
            "balance": float(a.balance), "currency": a.currency,
            "description": a.description, "is_active": a.is_active,
            "created_at": a.created_at.isoformat() if a.created_at else None
        }
        for a in accounts_result.scalars().all()
    ]

    # 2. 分类
    categories_result = await db.execute(
        select(Category).where(Category.user_id == user_id)
    )
    categories = [
        {
            "id": c.id, "name": c.name, "category_type": c.category_type,
            "description": c.description, "is_active": c.is_active,
            "parent_id": c.parent_id,
            "created_at": c.created_at.isoformat() if c.created_at else None
        }
        for c in categories_result.scalars().all()
    ]

    # 3. 标签
    tags_result = await db.execute(
        select(Tag).where(Tag.user_id == user_id)
    )
    tags = [
        {
            "id": t.id, "name": t.name, "color": t.color,
            "description": t.description, "is_active": t.is_active,
            "created_at": t.created_at.isoformat() if t.created_at else None
        }
        for t in tags_result.scalars().all()
    ]

    # 4. 交易
    transactions_result = await db.execute(
        select(Transaction).where(Transaction.user_id == user_id).order_by(Transaction.date)
    )
    transactions = []
    for t in transactions_result.scalars().all():
        # 获取关联标签
        tag_ids = [tag.id for tag in t.tags] if t.tags else []
        transactions.append({
            "id": t.id, "amount": float(t.amount), "description": t.description,
            "transaction_type": t.transaction_type,
            "category_id": t.category_id,
            "account_id": t.account_id,
            "from_account_id": t.from_account_id,
            "to_account_id": t.to_account_id,
            "date": t.date.isoformat() if t.date else None,
            "is_active": t.is_active,
            "tag_ids": tag_ids,
            "created_at": t.created_at.isoformat() if t.created_at else None
        })

    # 5. 预算
    budgets_result = await db.execute(
        select(Budget).where(Budget.user_id == user_id)
    )
    budgets = [
        {
            "id": b.id, "name": b.name, "category_id": b.category_id,
            "amount": float(b.amount),
            "period_start": b.period_start.isoformat() if b.period_start else None,
            "period_end": b.period_end.isoformat() if b.period_end else None,
            "description": b.description, "is_active": b.is_active,
            "created_at": b.created_at.isoformat() if b.created_at else None
        }
        for b in budgets_result.scalars().all()
    ]

    backup_data = {
        "version": "1.0",
        "exported_at": datetime.now().isoformat(),
        "user_id": user_id,
        "accounts": accounts,
        "categories": categories,
        "tags": tags,
        "transactions": transactions,
        "budgets": budgets,
    }

    # 保存为 JSON 文件
    filename = f"backup_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    filepath = os.path.join(BACKUP_DIR, filename)
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(backup_data, f, ensure_ascii=False, indent=2)

    return {
        "message": "备份创建成功",
        "filename": filename,
        "record_counts": {
            "accounts": len(accounts),
            "categories": len(categories),
            "tags": len(tags),
            "transactions": len(transactions),
            "budgets": len(budgets)
        }
    }


@router.get("/backup/list")
async def list_backups(
    user_id: int = Query(..., description="User ID")
):
    """列出用户的所有备份文件"""
    _ensure_backup_dir()

    backups = []
    if os.path.exists(BACKUP_DIR):
        for fname in sorted(os.listdir(BACKUP_DIR), reverse=True):
            if fname.startswith(f"backup_{user_id}_") and fname.endswith(".json"):
                fpath = os.path.join(BACKUP_DIR, fname)
                stat = os.stat(fpath)
                # 读取备份元信息
                try:
                    with open(fpath, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                    backups.append({
                        "filename": fname,
                        "size_bytes": stat.st_size,
                        "created_at": stat.st_ctime,
                        "exported_at": data.get("exported_at", ""),
                        "record_counts": {
                            "accounts": len(data.get("accounts", [])),
                            "categories": len(data.get("categories", [])),
                            "tags": len(data.get("tags", [])),
                            "transactions": len(data.get("transactions", [])),
                            "budgets": len(data.get("budgets", []))
                        }
                    })
                except Exception:
                    backups.append({
                        "filename": fname,
                        "size_bytes": stat.st_size,
                        "created_at": stat.st_ctime,
                        "exported_at": "",
                        "record_counts": {}
                    })

    return {"backups": backups}


@router.get("/backup/download/{filename}")
async def download_backup(
    filename: str,
    user_id: int = Query(..., description="User ID")
):
    """下载备份文件"""
    from fastapi import FileResponse as FR

    _ensure_backup_dir()

    # 安全检查：防止路径穿越
    safe_filename = os.path.basename(filename)
    if not safe_filename.startswith(f"backup_{user_id}_"):
        raise HTTPException(status_code=403, detail="无权访问此备份文件")

    filepath = os.path.join(BACKUP_DIR, safe_filename)
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="备份文件不存在")

    return FR(
        path=filepath,
        media_type="application/json",
        filename=safe_filename
    )


@router.delete("/backup/delete/{filename}")
async def delete_backup(
    filename: str,
    user_id: int = Query(..., description="User ID")
):
    """删除备份文件"""
    _ensure_backup_dir()

    safe_filename = os.path.basename(filename)
    if not safe_filename.startswith(f"backup_{user_id}_"):
        raise HTTPException(status_code=403, detail="无权删除此备份文件")

    filepath = os.path.join(BACKUP_DIR, safe_filename)
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="备份文件不存在")

    os.remove(filepath)
    return {"message": f"备份文件 {safe_filename} 已删除"}


@router.post("/restore")
async def restore_data(
    db: AsyncSession = Depends(get_db),
    backup_file: str = Query(..., description="备份文件名"),
    user_id: int = Query(..., description="User ID"),
    mode: str = Query("merge", description="恢复模式: merge(合并) 或 replace(替换)")
):
    """从备份文件恢复数据"""
    _ensure_backup_dir()

    safe_filename = os.path.basename(backup_file)
    if not safe_filename.startswith(f"backup_{user_id}_"):
        raise HTTPException(status_code=403, detail="无权访问此备份文件")

    filepath = os.path.join(BACKUP_DIR, safe_filename)
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="备份文件不存在")

    with open(filepath, 'r', encoding='utf-8') as f:
        backup_data = json.load(f)

    # 验证 user_id 匹配
    if backup_data.get("user_id") != user_id:
        raise HTTPException(status_code=403, detail="备份文件与当前用户不匹配")

    restored_counts = {"accounts": 0, "categories": 0, "tags": 0, "transactions": 0, "budgets": 0}

    try:
        if mode == "replace":
            # 替换模式：先删除用户现有数据（按依赖顺序）
            await db.execute(Transaction.__table__.delete().where(Transaction.user_id == user_id))
            await db.execute(Budget.__table__.delete().where(Budget.user_id == user_id))
            await db.execute(Tag.__table__.delete().where(Tag.user_id == user_id))
            await db.execute(Category.__table__.delete().where(Category.user_id == user_id))
            await db.execute(Account.__table__.delete().where(Account.user_id == user_id))
            await db.commit()

        # ID 映射（旧 ID → 新 ID）
        id_map = {"accounts": {}, "categories": {}, "tags": {}}

        # 1. 恢复分类
        for cat_data in backup_data.get("categories", []):
            cat_id = cat_data.pop("id", None)
            cat = Category(
                name=cat_data["name"],
                category_type=cat_data["category_type"],
                description=cat_data.get("description"),
                is_active=cat_data.get("is_active", True),
                parent_id=cat_data.get("parent_id"),
                user_id=user_id
            )
            db.add(cat)
            await db.flush()
            if cat_id:
                id_map["categories"][cat_id] = cat.id
            restored_counts["categories"] += 1

        # 2. 恢复标签
        for tag_data in backup_data.get("tags", []):
            tag_id = tag_data.pop("id", None)
            tag = Tag(
                name=tag_data["name"],
                color=tag_data.get("color"),
                description=tag_data.get("description"),
                is_active=tag_data.get("is_active", True),
                user_id=user_id
            )
            db.add(tag)
            await db.flush()
            if tag_id:
                id_map["tags"][tag_id] = tag.id
            restored_counts["tags"] += 1

        # 3. 恢复账户
        for acc_data in backup_data.get("accounts", []):
            acc_id = acc_data.pop("id", None)
            account = Account(
                name=acc_data["name"],
                account_type=acc_data["account_type"],
                balance=acc_data.get("balance", 0),
                currency=acc_data.get("currency", "CNY"),
                description=acc_data.get("description"),
                is_active=acc_data.get("is_active", True),
                user_id=user_id
            )
            db.add(account)
            await db.flush()
            if acc_id:
                id_map["accounts"][acc_id] = account.id
            restored_counts["accounts"] += 1

        # 4. 恢复交易
        for txn_data in backup_data.get("transactions", []):
            txn_data.pop("id", None)
            txn_data.pop("tag_ids", None)  # 单独处理标签关联
            txn_data.pop("created_at", None)

            # 映射分类 ID
            old_cat_id = txn_data.get("category_id")
            if old_cat_id and old_cat_id in id_map["categories"]:
                txn_data["category_id"] = id_map["categories"][old_cat_id]

            # 映射账户 ID
            old_acc_id = txn_data.get("account_id")
            if old_acc_id and old_acc_id in id_map["accounts"]:
                txn_data["account_id"] = id_map["accounts"][old_acc_id]

            old_from_id = txn_data.get("from_account_id")
            if old_from_id and old_from_id in id_map["accounts"]:
                txn_data["from_account_id"] = id_map["accounts"][old_from_id]

            old_to_id = txn_data.get("to_account_id")
            if old_to_id and old_to_id in id_map["accounts"]:
                txn_data["to_account_id"] = id_map["accounts"][old_to_id]

            txn_data["user_id"] = user_id
            txn = Transaction(**txn_data)
            db.add(txn)
            await db.flush()
            restored_counts["transactions"] += 1

        # 5. 恢复预算
        for bud_data in backup_data.get("budgets", []):
            bud_data.pop("id", None)
            bud_data.pop("created_at", None)

            old_cat_id = bud_data.get("category_id")
            if old_cat_id and old_cat_id in id_map["categories"]:
                bud_data["category_id"] = id_map["categories"][old_cat_id]

            bud_data["user_id"] = user_id
            bud_data["spent_amount"] = 0  # 重置已花费金额
            budget = Budget(**bud_data)
            db.add(budget)
            restored_counts["budgets"] += 1

        await db.commit()

        return {
            "message": "数据恢复成功",
            "mode": mode,
            "restored_counts": restored_counts
        }

    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"恢复失败: {str(e)}")


@router.post("/restore/upload")
async def restore_data_upload(
    db: AsyncSession = Depends(get_db),
    file: UploadFile = FileDep(..., description="备份 JSON 文件"),
    user_id: int = Query(..., description="User ID"),
    mode: str = Query("merge", description="恢复模式: merge(合并) 或 replace(替换)")
):
    """从上传的备份文件恢复数据"""
    _ensure_backup_dir()

    # 读取上传的文件
    content = await file.read()
    try:
        backup_data = json.loads(content)
    except Exception:
        raise HTTPException(status_code=400, detail="无效的 JSON 文件")

    # 验证
    if backup_data.get("user_id") != user_id:
        raise HTTPException(status_code=403, detail="备份文件与当前用户不匹配")

    # 先保存到备份目录
    safe_filename = f"backup_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_upload.json"
    filepath = os.path.join(BACKUP_DIR, safe_filename)
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(backup_data, f, ensure_ascii=False, indent=2)

    restored_counts = {"accounts": 0, "categories": 0, "tags": 0, "transactions": 0, "budgets": 0}

    try:
        if mode == "replace":
            # 删除现有数据（按依赖顺序）
            await db.execute(Transaction.__table__.delete().where(Transaction.user_id == user_id))
            await db.execute(Budget.__table__.delete().where(Budget.user_id == user_id))
            await db.execute(Tag.__table__.delete().where(Tag.user_id == user_id))
            await db.execute(Category.__table__.delete().where(Category.user_id == user_id))
            await db.execute(Account.__table__.delete().where(Account.user_id == user_id))
            await db.commit()

        # 恢复分类
        for cat_data in backup_data.get("categories", []):
            cat_data.pop("id", None)
            cat_data.pop("created_at", None)
            cat_data["user_id"] = user_id
            cat = Category(**cat_data)
            db.add(cat)
            await db.flush()
            restored_counts["categories"] += 1

        # 恢复标签
        for tag_data in backup_data.get("tags", []):
            tag_data.pop("id", None)
            tag_data.pop("created_at", None)
            tag_data["user_id"] = user_id
            tag = Tag(**tag_data)
            db.add(tag)
            await db.flush()
            restored_counts["tags"] += 1

        # 恢复账户
        for acc_data in backup_data.get("accounts", []):
            acc_data.pop("id", None)
            acc_data.pop("created_at", None)
            acc_data["user_id"] = user_id
            account = Account(**acc_data)
            db.add(account)
            await db.flush()
            restored_counts["accounts"] += 1

        # 恢复交易
        for txn_data in backup_data.get("transactions", []):
            txn_data.pop("id", None)
            txn_data.pop("tag_ids", None)
            txn_data.pop("created_at", None)
            txn_data["user_id"] = user_id
            txn = Transaction(**txn_data)
            db.add(txn)
            await db.flush()
            restored_counts["transactions"] += 1

        # 恢复预算
        for bud_data in backup_data.get("budgets", []):
            bud_data.pop("id", None)
            bud_data.pop("created_at", None)
            bud_data["user_id"] = user_id
            bud_data["spent_amount"] = 0
            budget = Budget(**bud_data)
            db.add(budget)
            restored_counts["budgets"] += 1

        await db.commit()

        return {
            "message": "数据恢复成功",
            "mode": mode,
            "restored_counts": restored_counts
        }

    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"恢复失败: {str(e)}")
